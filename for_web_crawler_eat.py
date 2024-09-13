import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

# 假设 country_domains 是一个包含所有国家名称及其域名后缀的字典
country_domains = {
    "Afghanistan": ".af",
    "Åland Islands": ".ax",
    "Albania": ".al",
    "Algeria": ".dz",
    "American Samoa": ".as",
    "Andorra": ".ad",
    "Angola": ".ao",
    "Anguilla": ".ai",
    "Antarctica": ".aq",
    "Antigua and Barbuda": ".ag",
    "Argentina": ".ar",
    "Armenia": ".am",
    "Aruba": ".aw",
    "Australia": ".au",
    "Austria": ".at",
    "Azerbaijan": ".az",
    "Bahamas": ".bs",
    "Bahrain": ".bh",
    "Bangladesh": ".bd",
    "Barbados": ".bb",
    "Belarus": ".by",
    "Belgium": ".be",
    "Belize": ".bz",
    "Benin": ".bj",
    "Bermuda": ".bm",
    "Bhutan": ".bt",
    "Bolivia": ".bo",
    "Bosnia and Herzegovina": ".ba",
    "Botswana": ".bw",
    "Bouvet Island": ".bv",
    "Brazil": ".br",
    "British Indian Ocean Territory": ".io",
    "Brunei Darussalam": ".bn",
    "Bulgaria": ".bg",
    "Burkina Faso": ".bf",
    "Burundi": ".bi",
    "Cambodia": ".kh",
    "Cameroon": ".cm",
    "Canada": ".ca",
    "Cape Verde": ".cv",
    "Cayman Islands": ".ky",
    "Central African Republic": ".cf",
    "Chad": ".td",
    "Chile": ".cl",
    "China": ".cn",
    "Christmas Island": ".cx",
    "Cocos (Keeling) Islands": ".cc",
    "Colombia": ".co",
    "Comoros": ".km",
    "Congo": ".cg",
    "Democratic Republic Of Congo": ".cd",
    "Cook Islands": ".ck",
    "Costa Rica": ".cr",
    "Côte d'Ivoire": ".ci",
    "Croatia": ".hr",
    "Cuba": ".cu",
    "Cyprus": ".cy",
    "Czech Republic": ".cz",
    "Denmark": ".dk",
    "Djibouti": ".dj",
    "Dominica": ".dm",
    "Dominican Republic": ".do",
    "Ecuador": ".ec",
    "Egypt": ".eg",
    "El Salvador": ".sv",
    "Equatorial Guinea": ".gq",
    "Eritrea": ".er",
    "Estonia": ".ee",
    "Ethiopia": ".et",
    "Falkland Islands (Malvinas)": ".fk",
    "Faroe Islands": ".fo",
    "Fiji": ".fj",
    "Finland": ".fi",
    "France": ".fr",
    "French Guiana": ".gf",
    "French Polynesia": ".pf",
    "French Southern Territories": ".tf",
    "Gabon": ".ga",
    "Gambia": ".gm",
    "Georgia": ".ge",
    "Germany": ".de",
    "Ghana": ".gh",
    "Gibraltar": ".gi",
    "Greece": ".gr",
    "Greenland": ".gl",
    "Grenada": ".gd",
    "Guadeloupe": ".gp",
    "Guam": ".gu",
    "Guatemala": ".gt",
    "Guernsey": ".gg",
    "Guinea": ".gn",
    "Guinea-Bissau": ".gw",
    "Guyana": ".gy",
    "Haiti": ".ht",
    "Heard Island and McDonald Islands": ".hm",
    "Holy See (Vatican City State)": ".va",
    "Honduras": ".hn",
    "Hong Kong": ".hk",
    "Hungary": ".hu",
    "Iceland": ".is",
    "India": ".in",
    "Indonesia": ".id",
    "Iran": ".ir",
    "Iraq": ".iq",
    "Ireland": ".ie",
    "Isle of Man": ".im",
    "Israel": ".il",
    "Italy": ".it",
    "Jamaica": ".jm",
    "Japan": ".jp",
    "Jersey": ".je",
    "Jordan": ".jo",
    "Kazakhstan": ".kz",
    "Kenya": ".ke",
    "Kiribati": ".ki",
    "Korea, Democratic People's Republic of": ".kp",
    "Korea": ".kr",
    "Kuwait": ".kw",
    "Kyrgyzstan": ".kg",
    "Lao People's Democratic Republic": ".la",
    "Latvia": ".lv",
    "Lebanon": ".lb",
    "Lesotho": ".ls",
    "Liberia": ".lr",
    "Libya": ".ly",
    "Liechtenstein": ".li",
    "Lithuania": ".lt",
    "Luxembourg": ".lu",
    "Macao": ".mo",
    "Macedonia": ".mk",
    "Madagascar": ".mg",
    "Malawi": ".mw",
    "Malaysia": ".my",
    "Maldives": ".mv",
    "Mali": ".ml",
    "Malta": ".mt",
    "Marshall Islands": ".mh",
    "Martinique": ".mq",
    "Mauritania": ".mr",
    "Mauritius": ".mu",
    "Mayotte": ".yt",
    "Mexico": ".mx",
    "Micronesia, Federated States of": ".fm",
    "Moldova": ".md",
    "Monaco": ".mc",
    "Mongolia": ".mn",
    "Montenegro": ".me",
    "Montserrat": ".ms",
    "Morocco": ".ma",
    "Mozambique": ".mz",
    "Myanmar": ".mm",
    "Namibia": ".na",
    "Nauru": ".nr",
    "Nepal": ".np",
    "Netherlands": ".nl",
    "New Caledonia": ".nc",
    "New Zealand": ".nz",
    "Nicaragua": ".ni",
    "Niger": ".ne",
    "Nigeria": ".ng",
    "Niue": ".nu",
    "Norfolk Island": ".nf",
    "Northern Mariana Islands": ".mp",
    "Norway": ".no",
    "Oman": ".om",
    "Pakistan": ".pk",
    "Palau": ".pw",
    "Palestinian Territory": ".ps",
    "Panama": ".pa",
    "Papua New Guinea": ".pg",
    "Paraguay": ".py",
    "Peru": ".pe",
    "Philippines": ".ph",
    "Pitcairn": ".pn",
    "Poland": ".pl",
    "Portugal": ".pt",
    "Puerto Rico": ".pr",
    "Qatar": ".qa",
    "Réunion": ".re",
    "Romania": ".ro",
    "Russia": ".ru",
    "Rwanda": ".rw",
    "Saint Barthélemy": ".bl",
    "Saint Helena": ".sh",
    "Saint Kitts and Nevis": ".kn",
    "Saint Lucia": ".lc",
    "Saint Martin (French part)": ".mf",
    "Saint Pierre and Miquelon": ".pm",
    "Saint Vincent and the Grenadines": ".vc",
    "Samoa": ".ws",
    "San Marino": ".sm",
    "Sao Tome and Principe": ".st",
    "Saudi Arabia": ".sa",
    "Senegal": ".sn",
    "Serbia": ".rs",
    "Seychelles": ".sc",
    "Sierra Leone": ".sl",
    "Singapore": ".sg",
    "Slovakia": ".sk",
    "Slovenia": ".si",
    "Solomon Islands": ".sb",
    "Somalia": ".so",
    "South Africa": ".za",
    "South Georgia and the South Sandwich Islands": ".gs",
    "Spain": ".es",
    "Sri Lanka": ".lk",
    "Sudan": ".sd",
    "Suriname": ".sr",
    "Svalbard and Jan Mayen": ".sj",
    "Swaziland": ".sz",
    "Sweden": ".se",
    "Switzerland": ".ch",
    "Syria": ".sy",
    "Taiwan, Province of China": ".tw",
    "Tajikistan": ".tj",
    "Tanzania": ".tz",
    "Thailand": ".th",
    "Timor-Leste": ".tl",
    "Togo": ".tg",
    "Tokelau": ".tk",
    "Tonga": ".to",
    "Trinidad and Tobago": ".tt",
    "Tunisia": ".tn",
    "Turkey": ".tr",
    "Turkmenistan": ".tm",
    "Turks and Caicos Islands": ".tc",
    "Tuvalu": ".tv",
    "Uganda": ".ug",
    "Ukraine": ".ua",
    "Uae": ".ae",
    "United Kingdom": ".uk",
    "United States": ".us",
    "United States Minor Outlying Islands": ".um",
    "Uruguay": ".uy",
    "Uzbekistan": ".uz",
    "Vanuatu": ".vu",
    "Venezuela, Bolivarian Republic of": ".ve",
    "Viet Nam": ".vn",
    "Virgin Islands, British": ".vg",
    "Virgin Islands, U.S.": ".vi",
    "Wallis and Futuna": ".wf",
    "Western Sahara": ".eh",
    "Yemen": ".ye",
    "Zambia": ".zm",
    "Zimbabwe": ".zw"
}

# 获取用户输入的国家和人名
countries_input = input("请输入国家名称（英文），多个国家用英文逗号隔开：")
person_name = input("请输入人名：")

# 将国家名称标准化为首字母大写
countries = [country.strip() for country in countries_input.split(",")]

# 检查国家是否在字典中
for country in countries:
    if country not in country_domains:
        raise ValueError(f"输入的国家名称 {country} 不在 country_domains 字典中")

# 指定要保存的Excel文件路径
excel_file_path = "C:\\Users\\yuesen\\Desktop\\爬虫参数.xlsx"

# 如果文件不存在，则创建新的工作簿和工作表
if not os.path.exists(excel_file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Crawling Parameters"
    # 设置表头
    headers = ["crawled", "course", "country", "pages", "file_name", "save_path", "start_url"]
    ws.append(headers)
else:
    # 如果文件存在，则加载工作簿和工作表
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active

# 获取当前最大 course 值，以确保新的 course 列继续递增
if ws.max_row > 1:
    max_course = ws.cell(row=ws.max_row, column=2).value
else:
    max_course = 0

# 生成每个国家的数据并追加到工作表
course_start = max_course + 1
file_names = ["经销商1", "进口商1", "代理商1", "经销商2", "进口商2", "代理商2"]

for country in countries:
    save_path = f"C:\\Users\\yuesen\\Desktop\\6+1\\{person_name}\\{country}"
    for i, file_name in enumerate(file_names):
        crawled = 0
        course = course_start
        course_start += 1
        pages = 300
        start_url_type = ""

        # 根据文件名生成 start_url 中的搜索词
        if "经销商" in file_name:
            start_url_type = "distributor"
            if "2" in file_name:
                start_url_type += f"+site%3A{country_domains[country]}"
        elif "进口商" in file_name:
            start_url_type = "importer"
            if "2" in file_name:
                start_url_type = f"import+site%3A{country_domains[country]}"
        elif "代理商" in file_name:
            start_url_type = "agent"
            if "2" in file_name:
                start_url_type += f"+site%3A{country_domains[country]}"

        start_url = f'https://cn.bing.com/search?q={country}+oxygen+concentrator+{start_url_type}&first=1&FORM=PERE&ensearch=1'

        # 将数据添加到工作表
        ws.append([crawled, course, country, pages, file_name, save_path, start_url])

# 设置列宽
for col in range(1, ws.max_column + 1):
    max_length = 0
    column = get_column_letter(col)
    for cell in ws[column]:
        if len(str(cell.value)) > max_length:
            max_length = len(cell.value)
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# 确保保存路径存在
for country in countries:
    country_save_path = f"C:\\Users\\yuesen\\Desktop\\6+1\\{person_name}\\{country}"
    os.makedirs(country_save_path, exist_ok=True)

# 保存Excel文件
wb.save(excel_file_path)

print(f"参数已成功追加写入Excel文件，文件保存路径为：{excel_file_path}")
